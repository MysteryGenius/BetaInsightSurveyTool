from __future__ import annotations

import re
import zipfile
from pathlib import Path
from typing import Iterator

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from surveytool.charts.errors import ErrorCode, SurveyToolError
from surveytool.core.hygiene import normalize_label
from surveytool.core.model import (
    CodeRole,
    Question,
    QuestionType,
    Response,
    ScaleCode,
    Survey,
)
from surveytool.core.nonsubstantive import classify_response, is_nonsubstantive_label
from surveytool.core.scale_library import identify_family, resolve_roles

_QID_RE = re.compile(r"\[([^\]]+)\]")
_SKIP_QIDS = frozenset({"record", "uuid", "uid"})


def _iter_until_blank(ws) -> Iterator[tuple]:
    """Yield rows, stopping as soon as every cell in a row is None."""
    for row in ws.iter_rows(values_only=True):
        if all(v is None for v in row):
            break
        yield row


def parse_codebook(ws_datamap) -> list[Question]:
    """
    Parse the Rakuten Datamap sheet into a list of Questions.

    The sheet has 3 columns:
      col 0: question header ('[QID]: text') or '' for code rows
      col 1: numeric code (for code rows) or None
      col 2: label text (for code rows) or None
    """
    questions: list[Question] = []
    current_qid: str | None = None
    current_text: str | None = None
    current_qtype_hint: str | None = None  # 'open_numeric' | 'open_text' | None
    current_codes: list[tuple[int, str]] = []  # (code, raw_label)

    def _flush() -> None:
        if current_qid is None:
            return
        if current_qid in _SKIP_QIDS:
            return
        _build_question(current_qid, current_text or "", current_qtype_hint, current_codes, questions)

    for row in ws_datamap.iter_rows(values_only=True):
        col0, col1, col2 = (row[i] if i < len(row) else None for i in range(3))

        # All-None rows terminate the sheet (or are blank separators between blocks)
        if col0 is None and col1 is None and col2 is None:
            continue

        if col0 and isinstance(col0, str) and col0.strip():
            cell = col0.strip()
            m = _QID_RE.match(cell)
            if m:
                # Flush previous block
                _flush()
                current_qid = m.group(1)
                current_text = cell[m.end():].lstrip(":").strip()
                current_qtype_hint = None
                current_codes = []
            elif cell.startswith("Open numeric"):
                current_qtype_hint = "open_numeric"
            elif cell.startswith("Open text"):
                current_qtype_hint = "open_text"
            # 'Values: N-M' rows — skip
        elif col1 is not None and col2 is not None:
            # Code row: ('', code, label)
            try:
                code = int(col1)
            except (TypeError, ValueError):
                continue
            current_codes.append((code, str(col2)))

    _flush()
    return questions


def _build_question(
    qid: str,
    text: str,
    qtype_hint: str | None,
    codes: list[tuple[int, str]],
    out: list[Question],
) -> None:
    if qtype_hint == "open_numeric":
        out.append(
            Question(
                qid=qid,
                text=text,
                qtype=QuestionType.numeric_open,
                base_eligible=False,
                is_demographic=qid.startswith("S"),
            )
        )
        return

    if qtype_hint == "open_text":
        out.append(
            Question(
                qid=qid,
                text=text,
                qtype=QuestionType.text_open,
                base_eligible=False,
                is_demographic=qid.startswith("S"),
            )
        )
        return

    if qid.startswith("S"):
        labels = [
            ScaleCode(code=code, label=label, role=CodeRole.excluded)
            for code, label in codes
        ]
        out.append(
            Question(
                qid=qid,
                text=text,
                qtype=QuestionType.demographic,
                labels=labels,
                base_eligible=False,
                is_demographic=True,
            )
        )
        return

    # Analysis question: try to resolve as scale
    if not codes:
        return

    norm_labels = [normalize_label(label) for _, label in codes]

    # Detect non-substantive labels before passing to scale library
    ns_overrides: dict[str, CodeRole] = {}
    for nl in norm_labels:
        if is_nonsubstantive_label(nl):
            ns_overrides[nl] = CodeRole.nonsubstantive

    # Substantive labels only go to the scale library
    substantive_norm = [nl for nl in norm_labels if nl not in ns_overrides]

    roles = resolve_roles(substantive_norm, override=ns_overrides)
    # resolve_roles only processes labels it receives; merge ns_overrides back
    for ns_label, ns_role in ns_overrides.items():
        roles[ns_label] = ns_role

    # Determine direction from the lowest-code substantive label's role
    substantive_codes_sorted = [
        (code, normalize_label(label))
        for code, label in sorted(codes, key=lambda x: x[0])
        if normalize_label(label) not in ns_overrides
    ]
    direction: str | None = None
    if substantive_codes_sorted:
        lowest_role = roles.get(substantive_codes_sorted[0][1])
        if lowest_role is CodeRole.bottom:
            direction = "ascending"
        elif lowest_role is CodeRole.top:
            direction = "descending"

    scale_codes = [
        ScaleCode(
            code=code,
            label=normalize_label(label),
            role=roles[normalize_label(label)],
            numeric_value=code,
        )
        for code, label in codes
    ]

    family = identify_family(
        [normalize_label(label) for _, label in codes if normalize_label(label) not in ns_overrides]
    )

    out.append(
        Question(
            qid=qid,
            text=text,
            qtype=QuestionType.scale,
            labels=scale_codes,
            scale_direction=direction,
            scale_family=family,
            base_eligible=True,
        )
    )


def load_respondents(
    ws_a1,
    questions: list[Question],
) -> tuple[list[Response], int]:
    """
    Load respondents from the A1 sheet.

    Returns (responses, n_raw) where n_raw is the number of data rows processed.
    Stops at the first all-None row.
    """
    rows = ws_a1.iter_rows(values_only=True)

    header = next(rows)
    col_index: dict[str, int] = {
        str(h): i for i, h in enumerate(header) if h is not None
    }

    flag_col: int | None = col_index.get("FLAG")
    uuid_col: int | None = col_index.get("uuid")
    record_col: int | None = col_index.get("record")

    qid_to_col: dict[str, int] = {}
    for q in questions:
        if q.qid in col_index:
            qid_to_col[q.qid] = col_index[q.qid]

    q_by_qid: dict[str, Question] = {q.qid: q for q in questions}

    responses: list[Response] = []
    n_raw = 0

    for row in rows:
        # Stop on all-None row
        if all(v is None for v in row):
            break

        # Determine respondent_id
        respondent_id: str
        if uuid_col is not None and row[uuid_col] is not None:
            respondent_id = str(row[uuid_col])
        elif record_col is not None and row[record_col] is not None:
            respondent_id = str(row[record_col])
        else:
            continue

        n_raw += 1

        vendor_flag: str | None = None
        if flag_col is not None and row[flag_col] is not None:
            vendor_flag = str(row[flag_col])

        for qid, col_idx in qid_to_col.items():
            raw = row[col_idx]
            # Coerce float with no fractional part to int (Excel sometimes gives 4.0)
            if isinstance(raw, float) and raw == int(raw):
                raw = int(raw)

            q = q_by_qid[qid]
            state = classify_response(raw, q)

            responses.append(
                Response(
                    respondent_id=respondent_id,
                    qid=qid,
                    raw_value=raw,
                    state=state,
                    vendor_flag=vendor_flag,
                )
            )

    return responses, n_raw


def load(path: Path, survey_id: str) -> Survey:
    """Load a Rakuten xlsx file and return a canonical Survey."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except (InvalidFileException, OSError, KeyError, zipfile.BadZipFile) as exc:
        raise SurveyToolError(
            ErrorCode.FILE_UNREADABLE,
            "This file could not be opened as a Rakuten export.",
            detail=str(exc),
            next_action="Confirm the file is a complete, un-corrupted .xlsx export, or pick a different file.",
        ) from exc

    try:
        try:
            ws_datamap = wb["Datamap"]
        except KeyError as exc:
            raise SurveyToolError(
                ErrorCode.MISSING_SHEET,
                "This file does not look like a Rakuten export.",
                detail="No Datamap sheet was found.",
                next_action="Check the vendor setting, or pick a different file.",
            ) from exc

        try:
            ws_respondents = wb["A1"]
        except KeyError as exc:
            raise SurveyToolError(
                ErrorCode.MISSING_SHEET,
                "This file does not look like a Rakuten export.",
                detail="No A1 respondent-data sheet was found.",
                next_action="Check the vendor setting, or pick a different file.",
            ) from exc

        questions = parse_codebook(ws_datamap)
        if not questions:
            raise SurveyToolError(
                ErrorCode.NO_QUESTIONS_FOUND,
                "This file was read but has no computable questions in it.",
                next_action="Check the Datamap sheet, or pick a different file.",
            )
        responses, n_raw = load_respondents(ws_respondents, questions)
    finally:
        wb.close()

    return Survey(
        id=survey_id,
        n_raw=n_raw,
        n_analysis=n_raw,
        questions=questions,
        responses=responses,
    )
