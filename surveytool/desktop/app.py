from __future__ import annotations

import csv
import logging
import tempfile
import uuid
import zipfile
from dataclasses import asdict, dataclass
from pathlib import Path

import openpyxl
import pandas as pd
from fastapi import FastAPI, Request, UploadFile, Form
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from openpyxl.utils.exceptions import InvalidFileException
from starlette.background import BackgroundTask

from surveytool.charts import render_charts
from surveytool.charts.chart_data import build_chart_data
from surveytool.charts.errors import ErrorCode, NoSessionError, SurveyToolError, WarningCode
from surveytool.charts.plotly_renderer import build_figure, rasterize
from surveytool.compute.breaks_compute import QuestionResult, compute_breaks_crosstab
from surveytool.compute.cross_tab import compute_cross_tab, detect_available_demographics
from surveytool.core.break_filter import validate_request
from surveytool.core.breaks_models import (
    CrosstabRequest,
    ExportRequest,
    ProjectionRequest,
    RegistryCategory,
    RegistryDimension,
    RegistryPayload,
)
from surveytool.core.breaks_payload import attach_chart_permissions
from surveytool.core.config import ToolConfig
from surveytool.core.demographic_registry import ResolvedRegistry, load_registry
from surveytool.core.model import Survey
from surveytool.core.projection import Projection, project
from surveytool.core.respondent_frame import to_respondent_frame
from surveytool.core.straightliner import detect_straightliners
from surveytool.findings.sheet import build_findings_sheet
from surveytool.ingest.milieu import _HEADER_RE as _MILIEU_HEADER_RE

_STATIC_DIR = Path(__file__).parent / "static"
_DEMOGRAPHICS_DIR = Path(__file__).parent.parent / "core" / "demographics"
_logger = logging.getLogger(__name__)

app = FastAPI()

_VENDORS = ("rakuten", "milieu", "toluna")
_VENDOR_LABELS = {"rakuten": "Rakuten", "milieu": "Milieu", "toluna": "Toluna"}

_CANONICAL_PATH = _DEMOGRAPHICS_DIR / "canonical.yaml"
_VENDOR_REGISTRY_PATHS = {
    vendor: _DEMOGRAPHICS_DIR / f"vendor_{vendor}.yaml" for vendor in _VENDORS
}


@app.exception_handler(SurveyToolError)
async def _survey_tool_error_handler(request: Request, exc: SurveyToolError) -> JSONResponse:
    if exc.code is ErrorCode.INTERNAL:
        _logger.exception("Internal error handling %s %s", request.method, request.url.path, exc_info=exc)
    return JSONResponse(status_code=exc.status_code, content=exc.to_response())


@app.exception_handler(Exception)
async def _unhandled_error_handler(request: Request, exc: Exception) -> JSONResponse:
    _logger.exception("Unhandled error handling %s %s", request.method, request.url.path, exc_info=exc)
    fallback = SurveyToolError(
        ErrorCode.INTERNAL,
        "Something failed while processing your request. Check the app log for details.",
    )
    return JSONResponse(status_code=fallback.status_code, content=fallback.to_response())


@dataclass
class Session:
    survey: Survey
    respondent_frame: pd.DataFrame
    findings: list
    vendor: str
    registry: ResolvedRegistry


def _default_config() -> ToolConfig:
    """Config for the breaks endpoints.

    The pre-existing endpoints let compute_cross_tab default its own config,
    so no session-level config is loaded at upload time. The breaks core
    functions take config as a required parameter, so this supplies the same
    defaults explicitly. A per-project config loader is a later concern.
    """
    return ToolConfig()


def _load_session_registry(vendor: str, survey: Survey) -> ResolvedRegistry:
    """Load and validate the demographic registry for an uploaded file.

    `surveys={vendor: survey}` runs the unmapped-source-value check (rule 2)
    against the file just loaded, so a source value this vendor's mapping
    doesn't cover fails the upload rather than silently disappearing from a
    later break. Any failure raises a SurveyToolError, which the app-level
    handler turns into a proper error response — the upload hard-fails and
    no partially-usable session is created.
    """
    return load_registry(_CANONICAL_PATH, _VENDOR_REGISTRY_PATHS, surveys={vendor: survey})


_SESSIONS: dict[str, Session] = {}


def get_session(session_id: str) -> Session | None:
    return _SESSIONS.get(session_id)


def _sniff_vendor(path: Path) -> str | None:
    """Identify which supported vendor a file's format matches, if any.

    Reuses each adapter's own distinguishing signal: the Datamap sheet for
    Rakuten and the RespondentData.Text sheet for Toluna (see the
    MISSING_SHEET checks in rakuten.py / toluna.py), and the bracketed
    header pattern for Milieu (milieu.py's _HEADER_RE). Reads only what's
    needed to identify the format, not a full parse. Returns None if the
    file matches no known vendor or can't be opened at all.
    """
    if path.suffix.lower() == ".xlsx":
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
        except (InvalidFileException, OSError, KeyError, zipfile.BadZipFile):
            return None
        try:
            sheetnames = set(wb.sheetnames)
        finally:
            wb.close()
        if "Datamap" in sheetnames:
            return "rakuten"
        if "RespondentData.Text" in sheetnames:
            return "toluna"
        return None

    try:
        with path.open(encoding="utf-8-sig", newline="") as f:
            header = next(csv.reader(f), None)
    except (OSError, UnicodeDecodeError, csv.Error):
        return None
    if header is not None and any(_MILIEU_HEADER_RE.match(cell.strip()) for cell in header):
        return "milieu"
    return None


def _load_survey(path: Path, survey_id: str, vendor: str) -> Survey:
    if vendor == "rakuten":
        from surveytool.ingest.rakuten import load
    elif vendor == "milieu":
        from surveytool.ingest.milieu import load
    elif vendor == "toluna":
        from surveytool.ingest.toluna import load
    else:
        raise SurveyToolError(
            ErrorCode.VENDOR_MISMATCH,
            "This vendor is not recognised.",
            detail=f"Unknown vendor {vendor!r}.",
            next_action=f"Choose one of: {', '.join(_VENDORS)}.",
        )

    detected = _sniff_vendor(path)
    if detected is not None and detected != vendor:
        raise SurveyToolError(
            ErrorCode.VENDOR_MISMATCH,
            f"This looks like a {_VENDOR_LABELS[detected]} export, but {_VENDOR_LABELS[vendor]} is selected.",
            detail=f"Detected vendor: {detected}. Selected vendor: {vendor}.",
            next_action="Change the vendor setting, or pick a different file.",
        )

    return load(path, survey_id)


def _require_session(session_id: str) -> Session:
    session = get_session(session_id)
    if session is None:
        raise NoSessionError()
    return session


@app.post("/api/upload")
async def upload(file: UploadFile, vendor: str = Form(...), survey_id: str = Form(...)):
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir) / Path(file.filename).name
        tmp_path.write_bytes(await file.read())

        survey = _load_survey(tmp_path, survey_id, vendor)
        registry = _load_session_registry(vendor, survey)
        excluded = detect_straightliners(survey.responses, survey.questions)
        findings = build_findings_sheet(survey, exclude_respondent_ids=excluded or None)

    respondent_frame = to_respondent_frame(survey)
    session_id = str(uuid.uuid4())
    _SESSIONS[session_id] = Session(
        survey=survey,
        respondent_frame=respondent_frame,
        findings=findings,
        vendor=vendor,
        registry=registry,
    )

    warnings = []
    if excluded:
        warnings.append({
            "code": WarningCode.STRAIGHTLINERS_EXCLUDED.value,
            "message": f"{len(excluded)} respondent(s) were excluded for straightlining.",
        })

    return {
        "session_id": session_id,
        "n_raw": survey.n_raw,
        "n_analysis": survey.n_analysis,
        "question_count": len(survey.questions),
        "warnings": warnings,
    }


@app.get("/api/session/{session_id}/charts")
async def get_charts(session_id: str):
    session = _require_session(session_id)
    chart_data = build_chart_data(session.findings, session.survey.id, session.survey.questions)
    return {"charts": [asdict(c) for c in chart_data]}


@app.post("/api/session/{session_id}/export")
async def export_charts(
    session_id: str,
    out_dir: str = Form(...),
    qid: str | None = Form(None),
    chart_type: str | None = Form(None),
    breakdown_variable: str | None = Form(None),
):
    session = _require_session(session_id)

    chart_filter = None
    if qid is not None or chart_type is not None or breakdown_variable is not None:
        chart_filter = (
            lambda spec: spec.qid == qid
            and spec.chart_type == chart_type
            and spec.breakdown_variable == breakdown_variable
        )

    try:
        entries = render_charts(
            session.findings,
            Path(out_dir),
            session.survey.id,
            session.survey.questions,
            chart_filter=chart_filter,
        )
    except OSError as exc:
        raise SurveyToolError(
            ErrorCode.INTERNAL,
            "The export could not be saved to the chosen location.",
            detail=str(exc),
            next_action="Choose a different export location and try again.",
        ) from exc

    if chart_filter is not None and not entries:
        raise SurveyToolError(
            ErrorCode.INTERNAL,
            "This chart could not be found in the current session.",
            next_action="Reload the page and try exporting again.",
        )

    return {"exported": [e.filename for e in entries], "out_dir": out_dir}


@app.get("/api/session/{session_id}/demographics")
async def get_demographics(session_id: str):
    session = _require_session(session_id)

    availability = detect_available_demographics(session.survey)
    demographics = [
        {"key": key, "label": q.text}
        for key, q in {**availability.standard, **availability.conditional}.items()
    ]
    questions = [
        {"qid": q.qid, "text": q.text}
        for q in session.survey.questions
        if q.base_eligible and not q.is_demographic
    ]
    return {"demographics": demographics, "questions": questions}


@app.get("/api/session/{session_id}/cross-tab")
async def get_cross_tab(session_id: str, qid: str, demographic: str, significance: bool = False):
    session = _require_session(session_id)

    question = next((q for q in session.survey.questions if q.qid == qid), None)
    if question is None:
        # Not user-triggerable through the picker (which only ever sends qids
        # from /demographics) — this only fires on a malformed request.
        raise SurveyToolError(
            ErrorCode.INTERNAL,
            "This question could not be found in the current session.",
            detail=f"Unknown question {qid!r}.",
            next_action="Reload the page and choose a question from the list.",
        )

    try:
        result = compute_cross_tab(
            session.respondent_frame, session.survey, question, demographic,
            significance=significance,
        )
    except NotImplementedError as exc:
        # Significance stays inert until the restructure (build plan section 6,
        # item A) — the GUI never sends this parameter, so this path is not
        # reachable through normal use. No code in the minimum set fits "a
        # feature flag that isn't wired up yet," so this keeps its pre-existing
        # bare 400 response rather than being forced into the taxonomy.
        return JSONResponse(status_code=400, content={"error": str(exc)})

    return {
        "qid": result.qid,
        "question_text": result.question_text,
        "demographic": result.demographic,
        "cells": [
            {
                "subgroup_code": c.subgroup_code,
                "subgroup_label": c.subgroup_label,
                "n": c.n,
                "status": c.status.value,
                "t2b": c.t2b,
                "b2b": c.b2b,
                "mean": c.mean,
            }
            for c in result.cells
        ],
    }


@app.get("/api/session/{session_id}/demographics/registry")
async def get_demographics_registry(session_id: str) -> RegistryPayload:
    """The canonical registry as resolved for this session's vendor.

    This is the UI's sole pill-list source (build plan section 12), so
    nothing is pruned: every canonical dimension is emitted with its full
    category list, including categories and dimensions this vendor does not
    map. `available` plus `source_column` record what this vendor actually
    declares, so the UI can disable rather than hide.

    Distinct from the pre-existing GET .../demographics, which serves the
    legacy single-break picker off detect_available_demographics and is
    unchanged.
    """
    session = _require_session(session_id)
    vendor_dims = session.registry.dimensions_for_vendor(session.vendor)

    dimensions = []
    for name, canonical_dim in session.registry.canonical.dimensions.items():
        dim_mapping = vendor_dims.get(name)
        dimensions.append(
            RegistryDimension(
                dimension=name,
                label=canonical_dim.label,
                multi_select=canonical_dim.multi_select,
                available=dim_mapping is not None,
                source_column=dim_mapping.source_column if dim_mapping else None,
                categories=[
                    RegistryCategory(
                        value=category.value,
                        label=category.label,
                        order=category.order,
                        non_response=category.non_response,
                    )
                    for category in canonical_dim.categories
                ],
            )
        )

    return RegistryPayload(vendor=session.vendor, dimensions=dimensions)


@app.post("/api/session/{session_id}/projection")
async def post_projection(session_id: str, request: ProjectionRequest) -> Projection:
    """The shape of a break/filter request, without computing any question.

    Returns 200 even for an invalid spec: project() reports validation
    failures in-band in Projection.errors rather than raising, because the
    UI re-runs this on every pill change and an in-progress selection is
    not an error condition (build plan section 6).
    """
    session = _require_session(session_id)
    return project(
        request.break_spec,
        request.filter_spec,
        session.survey,
        session.respondent_frame,
        session.registry,
        session.vendor,
        _default_config(),
    )


@app.post("/api/session/{session_id}/crosstab")
async def post_crosstab(session_id: str, request: CrosstabRequest) -> list[QuestionResult]:
    """N-dimensional cross-tab for the requested questions.

    Validates first: compute_breaks_crosstab expects an already-validated
    request and does not check the spec itself, so validate_request is
    called here directly. It raises on the first failure and the app-level
    SurveyToolError handler turns that into a proper 4xx — unlike the
    projection endpoint, an invalid spec here is a real error, since the
    caller has asked for figures that cannot be computed.
    """
    session = _require_session(session_id)
    config = _default_config()

    validate_request(
        request.break_spec,
        request.filter_spec,
        session.registry,
        session.survey,
        session.vendor,
        session.respondent_frame,
    )

    question_ids = request.question_ids
    if question_ids is None:
        # Same set the existing /demographics endpoint offers.
        question_ids = [
            q.qid for q in session.survey.questions
            if q.base_eligible and not q.is_demographic
        ]

    results = compute_breaks_crosstab(
        session.respondent_frame,
        session.survey,
        request.break_spec.dimensions,
        request.filter_spec.selections,
        question_ids,
        session.registry,
        session.vendor,
        config,
    )
    return attach_chart_permissions(results, config)


@app.post("/api/session/{session_id}/breaks-export")
async def post_breaks_export(session_id: str, request: ExportRequest):
    """Render one break/filter chart to a PNG via the Plotly/kaleido renderer
    (build plan section 11) and return the image file.

    Follows the same pipeline post_crosstab already uses (validate_request ->
    compute_breaks_crosstab -> attach_chart_permissions), scoped to the
    single requested question_id, then hands the resulting QuestionResult to
    plotly_renderer.build_figure. build_figure itself refuses (raising
    SurveyToolError/INTERNAL, mapped to a 500 by the app-level handler) a
    chart_type that isn't `permitted=True` in that result's permitted_charts
    -- this endpoint does not duplicate that check, since plotly_renderer is
    the layer that owns it (see plotly_renderer.build_figure's docstring).

    Routed at /breaks-export rather than /export because POST
    /api/session/{id}/export is already taken by the pre-existing
    matplotlib chart export, which must keep working unchanged.

    Returns the rendered PNG directly (FileResponse), rather than a
    manifest/out_dir pair like the legacy /export endpoint: ExportRequest
    carries no out_dir for this endpoint to write into, and a single chart
    is a single file, so there is no manifest to build.
    """
    session = _require_session(session_id)
    config = _default_config()

    validate_request(
        request.break_spec,
        request.filter_spec,
        session.registry,
        session.survey,
        session.vendor,
        session.respondent_frame,
    )

    results = compute_breaks_crosstab(
        session.respondent_frame,
        session.survey,
        request.break_spec.dimensions,
        request.filter_spec.selections,
        [request.question_id],
        session.registry,
        session.vendor,
        config,
    )
    results = attach_chart_permissions(results, config)
    question_result = results[0]

    fig = build_figure(
        question_result,
        request.chart_type,
        break_spec=request.break_spec.dimensions,
        filter_spec=request.filter_spec.selections,
    )

    out_path = Path(tempfile.gettempdir()) / f"breaks-export-{uuid.uuid4()}.png"
    rasterize(fig, out_path)

    return FileResponse(
        out_path,
        media_type="image/png",
        filename=f"{question_result.question_id}-{request.chart_type.value}.png",
        background=BackgroundTask(out_path.unlink, missing_ok=True),
    )


app.mount("/", StaticFiles(directory=_STATIC_DIR, html=True), name="static")
